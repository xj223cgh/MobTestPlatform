"""AI 生成用例 Excel 导出工具。

每次 AI 生成用例后，自动导出一份 Excel 文件到 workspace/outputs/excel/ 目录。
也可独立调用，按用例集导出。
"""
import logging
from datetime import datetime
from pathlib import Path
from typing import List

import yaml

logger = logging.getLogger(__name__)

_AI_ROOT = Path(__file__).resolve().parent
_CONFIG_FILE = _AI_ROOT / 'ai_config.yaml'
_DEFAULT_OUTPUT_DIR = _AI_ROOT / 'workspace' / 'outputs' / 'excel'

_config_cache: dict | None = None


def _load_excel_config() -> dict:
    """从 ai_config.yaml 加载 Excel 导出配置。"""
    global _config_cache
    if _config_cache is not None:
        return _config_cache
    try:
        with open(_CONFIG_FILE, 'r', encoding='utf-8') as f:
            cfg = yaml.safe_load(f) or {}
        _config_cache = cfg.get('output', {}).get('excel_export', {})
    except Exception:
        _config_cache = {}
    return _config_cache


def _get_columns() -> list:
    """获取列配置，优先从 ai_config.yaml 读取。"""
    cfg = _load_excel_config()
    columns = cfg.get('columns')
    if columns:
        return columns
    return [
        {'field': 'case_number',     'header': '用例编号', 'width': 18},
        {'field': 'case_name',       'header': '用例名称', 'width': 40},
        {'field': 'case_description','header': '用例描述', 'width': 50},
        {'field': 'priority',        'header': '优先级',   'width': 10},
        {'field': 'preconditions',   'header': '前置条件', 'width': 40},
        {'field': 'steps',           'header': '测试步骤', 'width': 60},
        {'field': 'expected_result', 'header': '预期结果', 'width': 50},
        {'field': 'test_data',       'header': '测试数据', 'width': 30},
    ]


def export_cases_to_excel(
    cases: List[dict],
    suite_name: str = '',
    output_dir: str | Path | None = None,
) -> Path | None:
    """将用例列表导出为 Excel 文件。

    Args:
        cases: 用例字典列表，字段见 ai_config.yaml 中 output.excel_export.columns。
        suite_name: 用例集名称，用于文件名。
        output_dir: 输出目录，默认为 workspace/outputs/excel/。

    Returns:
        导出成功返回文件路径，失败返回 None。
    """
    if not cases:
        logger.warning("No cases to export")
        return None

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        logger.error("openpyxl is not installed. Run: pip install openpyxl")
        return None

    out = Path(output_dir) if output_dir else _DEFAULT_OUTPUT_DIR
    out.mkdir(parents=True, exist_ok=True)

    cfg = _load_excel_config()
    sheet_name = cfg.get('sheet_name', '测试用例')
    columns = _get_columns()

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_font = Font(name='微软雅黑', size=10)
    cell_align = Alignment(vertical='top', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    for col_idx, col_cfg in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_cfg['header'])
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[cell.column_letter].width = col_cfg.get('width', 20)

    for row_idx, case in enumerate(cases, start=2):
        for col_idx, col_cfg in enumerate(columns, start=1):
            value = case.get(col_cfg['field'], '')
            if isinstance(value, list):
                value = '\n'.join(str(v) for v in value)
            cell = ws.cell(row=row_idx, column=col_idx, value=str(value) if value else '')
            cell.font = cell_font
            cell.alignment = cell_align
            cell.border = thin_border

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = 'A2'

    safe_name = ''.join(c for c in suite_name if c.isalnum() or c in ('_', '-', ' ')) or 'AI用例'
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"{safe_name}_{timestamp}.xlsx"
    filepath = out / filename

    try:
        wb.save(str(filepath))
        logger.info("Exported %d cases to %s", len(cases), filepath)
        return filepath
    except Exception as e:
        logger.error("Failed to save Excel: %s", e)
        return None
